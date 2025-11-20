from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active

ws['A1'] = "ID:"
ws['B1'] = "Name:"
ws['C1'] = "Second name:"

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 12

names = ["Tom", "Bob", "Steve", "King", "Lord", "Voldemort",
         "Harry", "Roan", "Allison", "Luke", "Paul"]

second_names = ["Green", "White", "Mclaren", "Black",
                "Burgers", "Rowling", "Dvorak"]

id_person = 1

for _ in range(2000):
    name = random.choice(names)
    second = random.choice(second_names)
    ws.append([id_person, name, second])
    id_person += 1

wb.save("products.xlsx")
